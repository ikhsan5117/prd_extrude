# -*- coding: utf-8 -*-
"""
Analisis Masterlist SPS Excel vs Database SpsItemLists
"""

import sys
import os

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# Install dependencies jika belum ada
def install_deps():
    import subprocess
    deps = ["openpyxl", "pyodbc", "tabulate"]
    for dep in deps:
        try:
            __import__(dep)
        except ImportError:
            print(f"Installing {dep}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", dep, "-q"])

install_deps()

import openpyxl
import pyodbc
from tabulate import tabulate

# ============================================================
# KONFIGURASI
# ============================================================
BASE_DIR = r"f:\extrude"

DB_SERVER   = "10.14.149.34,1433"
DB_NAME     = "prd_extrude_hose"
DB_USER     = "usrvelasto"
DB_PASSWORD = "H1s@na2025!!"

EXCEL_FILES = [
    {
        "file": "Masterlist SPS CHS 2 Layer DIG.xlsx",
        "label": "CHS 2 Layer DIG",
        "mode": "scan_all",
    },
    {
        "file": "Masterlist SPS CHS 3 Layer DIG.xlsx",
        "label": "CHS 3 Layer DIG",
        "mode": "scan_all",
    },
    {
        "file": "Masterlist SPS Double Layer_Digitalisasi.xlsx",
        "label": "Double Layer Digitalisasi",
        "mode": "scan_all",
    },
    {
        "file": "Masterlist SPS Double Layer.xlsx",
        "label": "Double Layer",
        "mode": "scan_all",
    },
]

# ============================================================
# FUNGSI HELPER
# ============================================================

def cell_val(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()

def scan_workbook(wb, label):
    """Scan semua sheet, cari kolom yang mengandung No.Document dan Item."""
    all_items = []
    keywords_nodoc = ["no. doc", "nodoc", "no.doc", "document", "no doc", "no.document", "no document"]
    keywords_item  = ["item list", "itemlist", "item", "part no", "part number"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: '{sheet_name}' | Rows: {ws.max_row} | Cols: {ws.max_column}")

        # Cari baris header di 5 baris pertama
        nodoc_col = None
        item_col  = None
        hrow_found = None

        for hrow in range(1, min(8, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                val_lower = cell_val(ws, hrow, col).lower()
                if nodoc_col is None and any(k in val_lower for k in keywords_nodoc):
                    nodoc_col  = col
                    hrow_found = hrow
                if item_col is None and any(k in val_lower for k in keywords_item):
                    item_col   = col
                    if hrow_found is None:
                        hrow_found = hrow

        if nodoc_col or item_col:
            print(f"    -> Header row: {hrow_found} | NoDoc col: {nodoc_col} | Item col: {item_col}")
            nc_lbl = f"Col {nodoc_col}({openpyxl.utils.get_column_letter(nodoc_col)}): '{cell_val(ws, hrow_found, nodoc_col)}'" if nodoc_col else "N/A"
            ic_lbl = f"Col {item_col}({openpyxl.utils.get_column_letter(item_col)}): '{cell_val(ws, hrow_found, item_col)}'"  if item_col  else "N/A"
            print(f"       NoDoc -> {nc_lbl}")
            print(f"       Item  -> {ic_lbl}")

            count_empty = 0
            sheet_items = 0
            for row in range(hrow_found + 1, ws.max_row + 1):
                nodoc = cell_val(ws, row, nodoc_col) if nodoc_col else ""
                item  = cell_val(ws, row, item_col)  if item_col  else ""

                if not nodoc and not item:
                    count_empty += 1
                    if count_empty > 15:
                        break
                    continue
                count_empty = 0

                if nodoc or item:
                    all_items.append({
                        "nodoc":  nodoc,
                        "item":   item,
                        "source": f"{label} [{sheet_name}]"
                    })
                    sheet_items += 1

            print(f"    -> Extracted: {sheet_items} items")
        else:
            # Tampilkan header row 1-3 untuk investigasi
            print(f"    -> Kolom NoDoc/Item tidak ditemukan, sample header:")
            for hrow in range(1, min(4, ws.max_row + 1)):
                vals = [cell_val(ws, hrow, c) for c in range(1, min(ws.max_column + 1, 20))]
                non_empty = [v for v in vals if v]
                if non_empty:
                    print(f"       Row {hrow}: {non_empty}")

    return all_items


# ============================================================
# MAIN
# ============================================================

def main():
    print()
    print("=" * 70)
    print("  ANALISIS MASTERLIST SPS EXCEL vs DATABASE SpsItemLists")
    print("=" * 70)

    # ----------------------------------------------------------
    # STEP 1: Baca semua Excel
    # ----------------------------------------------------------
    all_excel_items = []

    for cfg in EXCEL_FILES:
        filepath = os.path.join(BASE_DIR, cfg["file"])
        print(f"\n[+] Membuka: {cfg['file']}")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [ERROR] Gagal buka file: {e}")
            continue

        items = scan_workbook(wb, cfg["label"])
        all_excel_items.extend(items)
        wb.close()

    # ----------------------------------------------------------
    # STEP 2: Ringkasan dari Excel
    # ----------------------------------------------------------
    print()
    print("=" * 70)
    print("  RINGKASAN DATA DARI EXCEL")
    print("=" * 70)

    excel_pairs   = set()
    excel_items   = set()
    excel_nodocs  = set()

    for entry in all_excel_items:
        item  = entry["item"].strip()
        nodoc = entry["nodoc"].strip()
        if item or nodoc:
            excel_pairs.add((item, nodoc))
            if item:  excel_items.add(item)
            if nodoc: excel_nodocs.add(nodoc)

    print(f"\nTotal baris data dari Excel   : {len(all_excel_items)}")
    print(f"Unique pasangan (Item, NoDoc) : {len(excel_pairs)}")
    print(f"Unique Item List              : {len(excel_items)}")
    print(f"Unique No. Document           : {len(excel_nodocs)}")

    if not excel_items and not excel_nodocs:
        print("\n[WARN] Tidak ada data Item/NoDoc yang berhasil diekstrak!")
        print("  -> Silakan cek output 'sample header' di atas untuk investigasi manual.")
        return

    # ----------------------------------------------------------
    # STEP 3: Koneksi Database
    # ----------------------------------------------------------
    print()
    print("=" * 70)
    print(f"  KONEKSI DATABASE: {DB_NAME}")
    print("=" * 70)

    conn = None
    for driver in ["ODBC Driver 17 for SQL Server", "ODBC Driver 18 for SQL Server", "SQL Server"]:
        try:
            cs = (
                f"DRIVER={{{driver}}};"
                f"SERVER={DB_SERVER};"
                f"DATABASE={DB_NAME};"
                f"UID={DB_USER};"
                f"PWD={DB_PASSWORD};"
                f"TrustServerCertificate=yes;"
                f"Encrypt=no;"
            )
            conn = pyodbc.connect(cs, timeout=15)
            print(f"  [OK] Koneksi berhasil dengan driver: {driver}")
            break
        except Exception as e:
            print(f"  [--] Driver '{driver}' gagal: {e}")

    if conn is None:
        print("\n[ERROR] Semua driver gagal. Cek koneksi network/VPN ke 10.14.149.34")
        return

    cursor = conn.cursor()

    # ----------------------------------------------------------
    # STEP 4: Ambil data DB
    # ----------------------------------------------------------
    print("\nMengambil data dari tabel SpsItemLists...")
    try:
        cursor.execute("SELECT Id, ItemList, DocumentNumber FROM SpsItemLists ORDER BY DocumentNumber, ItemList")
        db_rows = cursor.fetchall()
    except Exception as e:
        print(f"[ERROR] Query gagal: {e}")
        conn.close()
        return

    print(f"  [OK] Total data di DB: {len(db_rows)} baris")

    db_pairs  = set()
    db_items  = set()
    db_nodocs = set()
    db_id_map = {}

    for row in db_rows:
        db_id, item, nodoc = row
        item  = (item  or "").strip()
        nodoc = (nodoc or "").strip()
        db_pairs.add((item, nodoc))
        db_items.add(item)
        db_nodocs.add(nodoc)
        db_id_map[(item, nodoc)] = db_id

    print(f"       Unique Item         : {len(db_items)}")
    print(f"       Unique No. Document : {len(db_nodocs)}")

    # ----------------------------------------------------------
    # STEP 5: Perbandingan
    # ----------------------------------------------------------
    missing_in_db = excel_pairs - db_pairs
    extra_in_db   = db_pairs   - excel_pairs
    matched       = excel_pairs & db_pairs

    print()
    print("=" * 70)
    print("  HASIL PERBANDINGAN EXCEL vs DATABASE")
    print("=" * 70)
    print(f"\n  [OK]   Cocok (ada di keduanya)             : {len(matched)}")
    print(f"  [MISS] KURANG di DB (Excel ada, DB tidak)  : {len(missing_in_db)}")
    print(f"  [XTRA] Extra di DB  (DB ada, Excel tidak)  : {len(extra_in_db)}")

    # ----------------------------------------------------------
    # STEP 6: Detail yang kurang
    # ----------------------------------------------------------
    if missing_in_db:
        print()
        print("=" * 70)
        print(f"  DATA YANG PERLU DITAMBAHKAN ({len(missing_in_db)} record)")
        print("=" * 70)

        missing_sorted = sorted(missing_in_db, key=lambda x: (x[1], x[0]))

        table_data = []
        for item, nodoc in missing_sorted:
            srcs = set(e["source"] for e in all_excel_items
                       if e["item"].strip() == item and e["nodoc"].strip() == nodoc)
            src_str = "; ".join(sorted(srcs))[:60]
            table_data.append([item or "(kosong)", nodoc or "(kosong)", src_str])

        print(tabulate(table_data,
                      headers=["Item List", "No. Document", "Sumber File"],
                      tablefmt="grid",
                      maxcolwidths=[45, 35, 45]))

        # Simpan CSV
        os.makedirs(os.path.join(BASE_DIR, "scratch"), exist_ok=True)
        csv_path = os.path.join(BASE_DIR, "scratch", "missing_sps_items.csv")
        with open(csv_path, "w", encoding="utf-8-sig") as f:
            f.write("ItemList,DocumentNumber,SumberFile\n")
            for item, nodoc in missing_sorted:
                srcs = set(e["source"] for e in all_excel_items
                           if e["item"].strip() == item and e["nodoc"].strip() == nodoc)
                src_str = "; ".join(sorted(srcs))
                f.write(f'"{item}","{nodoc}","{src_str}"\n')
        print(f"\n  [SAVED] CSV -> {csv_path}")

        # Simpan SQL INSERT
        sql_path = os.path.join(BASE_DIR, "scratch", "insert_missing_sps_items.sql")
        with open(sql_path, "w", encoding="utf-8") as f:
            f.write("-- SQL INSERT untuk data yang KURANG di SpsItemLists\n")
            f.write(f"-- Total: {len(missing_sorted)} records\n\n")
            f.write("USE prd_extrude_hose;\nGO\n\nBEGIN TRANSACTION;\n\n")
            for item, nodoc in missing_sorted:
                f.write(
                    f"INSERT INTO SpsItemLists (ItemList, DocumentNumber) "
                    f"VALUES (N'{item.replace(chr(39), chr(39)*2)}', "
                    f"N'{nodoc.replace(chr(39), chr(39)*2)}');\n"
                )
            f.write("\n-- Cek jumlah setelah insert:\n")
            f.write("-- SELECT COUNT(*) FROM SpsItemLists;\n")
            f.write("COMMIT;\n")
        print(f"  [SAVED] SQL -> {sql_path}")

    else:
        print("\n  [OK] Semua data Excel sudah ada di database SpsItemLists!")

    # ----------------------------------------------------------
    # STEP 7: Extra di DB
    # ----------------------------------------------------------
    if extra_in_db:
        extra_sorted = sorted(extra_in_db, key=lambda x: (x[1], x[0]))
        print()
        print("=" * 70)
        print(f"  DATA DI DB YANG TIDAK ADA DI EXCEL ({len(extra_in_db)} record)")
        print("=" * 70)
        tbl = []
        for item, nodoc in extra_sorted[:100]:
            tbl.append([db_id_map.get((item, nodoc), "?"), item or "(kosong)", nodoc or "(kosong)"])
        print(tabulate(tbl, headers=["DB Id", "Item List", "No. Document"],
                      tablefmt="simple", maxcolwidths=[8, 45, 35]))
        if len(extra_in_db) > 100:
            print(f"  ... dan {len(extra_in_db) - 100} record lagi")

    # ----------------------------------------------------------
    # STEP 8: NoDoc unik yang hilang
    # ----------------------------------------------------------
    missing_nodocs = excel_nodocs - db_nodocs
    if missing_nodocs:
        print()
        print("=" * 70)
        print(f"  No. DOCUMENT DARI EXCEL YANG BELUM ADA DI DB ({len(missing_nodocs)} dokumen)")
        print("=" * 70)
        for nd in sorted(missing_nodocs):
            print(f"  - {nd}")

    cursor.close()
    conn.close()

    print()
    print("=" * 70)
    print("  SELESAI!")
    print("=" * 70)
    print()


if __name__ == "__main__":
    main()
