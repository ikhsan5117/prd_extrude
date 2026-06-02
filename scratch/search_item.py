# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, pyodbc

BASE_DIR = r"f:\extrude"
SEARCH   = "TA2712"

FILES = [
    ("Masterlist SPS CHS 2 Layer DIG.xlsx",          ["Parameter Setting"]),
    ("Masterlist SPS CHS 3 Layer DIG.xlsx",           ["Master 1"]),
    ("Masterlist SPS Double Layer_Digitalisasi.xlsx", ["Parameter Setting", "Digitalisasi", "Print"]),
    ("Masterlist SPS Double Layer.xlsx",              ["Parameter Setting", "Print"]),
]

print(f"=== PENCARIAN: '{SEARCH}' (+ Machine) ===\n")

# ── EXCEL ────────────────────────────────────────────────────
print("[ EXCEL ]")
found_excel = False
seen = set()  # hindari duplikat nodoc+machine

for fname, sheets in FILES:
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, fname), read_only=True, data_only=True)
    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]

        # Temukan kolom Machine dan Item dari header row 3
        machine_col = None
        item_col    = None
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=3, column=col).value or "").strip().lower()
            if h == "machine"  and machine_col is None: machine_col = col
            if "item" in h     and item_col    is None: item_col    = col

        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value else ""
                if SEARCH.lower() in val.lower():
                    nodoc   = ws.cell(row=cell.row, column=4).value or ""
                    machine = ws.cell(row=cell.row, column=machine_col).value if machine_col else "?"
                    nodoc   = str(nodoc).strip()
                    machine = str(machine).strip()

                    key = (fname, sname, nodoc, machine)
                    if key in seen:
                        continue
                    seen.add(key)

                    print(f"  File    : {fname}")
                    print(f"  Sheet   : {sname}")
                    print(f"  Machine : {machine}")
                    print(f"  No.Doc  : {nodoc}")
                    print(f"  Item    : {val}")
                    print()
                    found_excel = True
    wb.close()

if not found_excel:
    print("  -> TIDAK DITEMUKAN di Excel\n")

# ── DATABASE ─────────────────────────────────────────────────
print("[ DATABASE - SpsItemLists ]")
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=10.14.149.34,1433;DATABASE=prd_extrude_hose;"
    "UID=usrvelasto;PWD=H1s@na2025!!;"
    "TrustServerCertificate=yes;Encrypt=no;", timeout=10
)
cur = conn.cursor()
cur.execute(
    "SELECT s.Id, s.ItemList, s.DocumentNumber, n.Machine "
    "FROM SpsItemLists s "
    "LEFT JOIN SpsNoDocs n ON s.DocumentNumber = n.DocumentNumber "
    "WHERE s.ItemList LIKE ? "
    "ORDER BY s.DocumentNumber",
    ("%" + SEARCH + "%",)
)
rows = cur.fetchall()
if rows:
    for r in rows:
        print(f"  Id={r[0]}  |  Machine={r[3]}  |  NoDoc={r[2]}  |  ItemList={r[1]}")
else:
    print(f"  -> '{SEARCH}' TIDAK DITEMUKAN di SpsItemLists")

cur.close()
conn.close()
print("\n=== SELESAI ===")
