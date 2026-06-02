# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE_DIR = r"f:\extrude"
SEARCH_TERMS = ["NA3330", "NA1840", "NA2720"]

FILES = [
    ("Masterlist SPS CHS 2 Layer DIG.xlsx",          ["Parameter Setting"]),
    ("Masterlist SPS CHS 3 Layer DIG.xlsx",           ["Master 1"]),
    ("Masterlist SPS Double Layer_Digitalisasi.xlsx", ["Parameter Setting", "Digitalisasi", "Print"]),
    ("Masterlist SPS Double Layer.xlsx",              ["Parameter Setting", "Print"]),
]

print(f"=== PENCARIAN BANYAK ITEM ===")

for term in SEARCH_TERMS:
    print(f"\n>> Mencari: {term}")
    found = False
    
    for fname, sheets in FILES:
        wb = openpyxl.load_workbook(os.path.join(BASE_DIR, fname), read_only=True, data_only=True)
        for sname in sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            
            # Cari kolom
            machine_col = None
            rev_col = None
            revdate_col = None
            
            for col in range(1, ws.max_column + 1):
                h = str(ws.cell(row=3, column=col).value or "").strip().lower()
                if "machine" in h and machine_col is None: machine_col = col
                if "rev." in h or "revision" in h:
                    if "date" in h: revdate_col = col
                    elif rev_col is None: rev_col = col
            
            # rev_col fallback
            if not rev_col: rev_col = 5  # biasanya col E (No. Rev)
            if not revdate_col: revdate_col = 7 # biasanya col G (Rev Date)
            
            for row in ws.iter_rows(min_row=4):
                for cell in row:
                    val = str(cell.value) if cell.value else ""
                    if term.lower() in val.lower():
                        nodoc   = ws.cell(row=cell.row, column=4).value or ""
                        machine = ws.cell(row=cell.row, column=machine_col).value if machine_col else "?"
                        rev     = ws.cell(row=cell.row, column=rev_col).value or ""
                        revdate = ws.cell(row=cell.row, column=revdate_col).value or ""
                        
                        print(f"  File   : {fname} [{sname}]")
                        print(f"  Machine: {machine}")
                        print(f"  No.Doc : {nodoc}")
                        print(f"  Revisi : Ke-{rev} (Tgl: {revdate})")
                        print(f"  Item   : {val}")
                        print()
                        found = True
        wb.close()
    
    if not found:
        print("  -> TIDAK DITEMUKAN di Excel\n")

print("\n=== SELESAI ===")
